import os
from openpyxl import load_workbook

wb = load_workbook(filename=f"input/example.xlsx")
sheet_names = wb.sheetnames
ws = wb[sheet_names[0]]

for row_num, row_val in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
    Name = row_val[0].value
    Age = row_val[1].value
    City = row_val[2].value
